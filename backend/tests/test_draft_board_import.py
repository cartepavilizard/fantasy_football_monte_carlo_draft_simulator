# -*- coding: utf-8 -*-
"""
Tests for the paper draft-board importer.

Synthetic .xlsx fixtures are built with openpyxl into pytest's tmp_path
(no dependency on any file outside the repo, no database required). The
pure functions (parse_board, overall_pick_for, normalize_name,
match_board_to_picks) are covered directly; apply_board is exercised with
a fake engine object so no Mongo is needed.
"""
import asyncio

import openpyxl
import pytest

from data_sources.draft_board_import import (
    BoardMatchResult,
    DraftBoard,
    MIN_MATCH_RATE,
    MIN_OWNER_OVERLAP,
    apply_board,
    match_board_to_picks,
    normalize_name,
    overall_pick_for,
    parse_board,
)
from models.sources import HistoricalPick


# ---------------------------------------------------------------------------
# xlsx fixture builder
# ---------------------------------------------------------------------------


def _write_board(tmp_path, team_names, rounds, leading_blanks=0):
    """
    Build a synthetic .xlsx board.

    rounds: dict[round_num] -> list[cell_text in slot order (slot 1..N)]
    leading_blanks: number of fully-blank rows before the header
    """
    path = tmp_path / "board.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Board"
    row = 1
    for _ in range(leading_blanks):
        row += 1
    # header
    ws.cell(row=row, column=1, value="Round")
    for i, name in enumerate(team_names):
        ws.cell(row=row, column=2 + i, value=name)
    row += 1
    for rn in sorted(rounds):
        ws.cell(row=row, column=1, value=f"Round #{rn}")
        for slot, text in enumerate(rounds[rn]):
            if text is not None:
                ws.cell(row=row, column=2 + slot, value=text)
        row += 1
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# overall_pick_for
# ---------------------------------------------------------------------------


def test_overall_pick_for_snake_rounds_1_to_3_yields_1_to_36():
    picks = []
    for rn in (1, 2, 3):
        for slot in range(1, 13):
            picks.append(overall_pick_for(rn, slot, team_count=12, snake=True))
    assert sorted(picks) == list(range(1, 37))
    assert len(set(picks)) == 36  # no duplicates


def test_overall_pick_for_snake_round_2_reversed():
    assert overall_pick_for(2, 12, 12, snake=True) == 13
    assert overall_pick_for(2, 1, 12, snake=True) == 24
    # round 1 is straight
    assert overall_pick_for(1, 1, 12, snake=True) == 1
    assert overall_pick_for(1, 12, 12, snake=True) == 12


def test_overall_pick_for_non_snake_is_straight():
    for rn in (1, 2, 3):
        for slot in range(1, 13):
            assert overall_pick_for(rn, slot, 12, snake=False) == (
                (rn - 1) * 12 + slot
            )


# ---------------------------------------------------------------------------
# normalize_name
# ---------------------------------------------------------------------------


def test_normalize_name_basics():
    assert normalize_name("Christian McCaffrey") == "christian mccaffrey"
    assert normalize_name("A.J. Brown") == "aj brown"
    assert normalize_name("Ja'Marr Chase") == "jamarr chase"
    assert normalize_name("Kenneth Walker III") == "kenneth walker"
    assert normalize_name("Bijan Robinson") == "bijan robinson"


def test_normalize_name_strips_trailing_tags_and_parentheticals():
    assert normalize_name("Davante Adams (LV)") == "davante adams"
    assert normalize_name("Christian McCaffrey RB") == "christian mccaffrey"
    assert normalize_name("Dallas Cowboys D/ST") == "dallas cowboys"


# ---------------------------------------------------------------------------
# parse_board
# ---------------------------------------------------------------------------


def test_parse_board_handles_blanks_variants_and_empty_cells(tmp_path):
    rounds = {
        1: ["Alpha", "Bravo", "Charlie", "Delta", "Echo", "Foxtrot"],
        2: ["Golf", None, "Hotel", "India", None, "Juliet"],  # blanks
        3: ["Kilo", "Lima", "Mike", "November", "Oscar", "Papa"],
    }
    path = _write_board(
        tmp_path,
        team_names=["T1", "T2", "T3", "T4", "T5", "T6"],
        rounds=rounds,
        leading_blanks=2,
    )
    board = parse_board(str(path))
    assert board.team_names == ["T1", "T2", "T3", "T4", "T5", "T6"]
    assert board.cells[(1, 1)] == "Alpha"
    assert board.cells[(1, 6)] == "Foxtrot"
    # blank cells absent
    assert (2, 2) not in board.cells
    assert (2, 5) not in board.cells
    assert board.cells[(2, 3)] == "Hotel"
    assert board.cells[(3, 6)] == "Papa"


def test_parse_board_accepts_label_variants(tmp_path):
    """'Round #1', 'Round 2', bare '3' all count as round rows."""
    path = tmp_path / "variants.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="Round")
    for i, name in enumerate(["A", "B", "C", "D"]):
        ws.cell(row=1, column=2 + i, value=name)
    ws.cell(row=2, column=1, value="Round #1")
    ws.cell(row=2, column=2, value="p1")
    ws.cell(row=3, column=1, value="Round 2")
    ws.cell(row=3, column=3, value="p2")
    ws.cell(row=4, column=1, value="3")  # bare integer
    ws.cell(row=4, column=4, value="p3")
    wb.save(path)
    board = parse_board(str(path))
    assert board.cells[(1, 1)] == "p1"
    assert board.cells[(2, 2)] == "p2"
    assert board.cells[(3, 3)] == "p3"


# ---------------------------------------------------------------------------
# helpers for match tests
# ---------------------------------------------------------------------------


def _pick(overall, round_num, guid, name, display=None, team_id=1):
    return HistoricalPick(
        espn_league_id=1,
        season=2024,
        overall_pick=overall,
        round_num=round_num,
        round_pick=overall,
        member_guid=guid,
        owner_display_name=display or guid,
        espn_team_id=team_id,
        raw_player_name=name,
        canonical_name=name,
        position="rb",
        is_keeper=False,
        bid_amount=None,
        historical_adp=None,
        draft_order_verified=False,  # fabricated ESPN picks awaiting re-order
    )




# ---------------------------------------------------------------------------
# owner-map derivation (4-team boards; parse_board requires >=4 columns)
# ---------------------------------------------------------------------------


def _four_owner_picks():
    """4 owners x 4 picks each. Board column order in tests is shuffled
    relative to ESPN team_id order to prove the owner map derives from
    roster overlap, not header names."""
    return [
        _pick(1, 1, "G1", "Christian McCaffrey", display="Alpha", team_id=1),
        _pick(2, 1, "G1", "Jahmyr Gibbs", display="Alpha", team_id=1),
        _pick(3, 1, "G1", "Garrett Wilson", display="Alpha", team_id=1),
        _pick(4, 1, "G1", "Breece Hall", display="Alpha", team_id=1),
        _pick(5, 1, "G2", "Bijan Robinson", display="Beta", team_id=2),
        _pick(6, 1, "G2", "Puka Nacua", display="Beta", team_id=2),
        _pick(7, 1, "G2", "Marvin Harrison", display="Beta", team_id=2),
        _pick(8, 1, "G2", "De'Von Achane", display="Beta", team_id=2),
        _pick(9, 1, "G3", "Tyreek Hill", display="Gamma", team_id=3),
        _pick(10, 1, "G3", "Najee Harris", display="Gamma", team_id=3),
        _pick(11, 1, "G3", "Drake London", display="Gamma", team_id=3),
        _pick(12, 1, "G3", "Kyle Pitts", display="Gamma", team_id=3),
        _pick(13, 1, "G4", "Saquon Barkley", display="Delta", team_id=4),
        _pick(14, 1, "G4", "Josh Jacobs", display="Delta", team_id=4),
        _pick(15, 1, "G4", "Amari Cooper", display="Delta", team_id=4),
        _pick(16, 1, "G4", "Trevor Lawrence", display="Delta", team_id=4),
    ]


def _board_from_columns(tmp_path, team_names, columns_by_slot):
    """Build a board where each slot's column is one owner's roster
    across rounds 1..N. columns_by_slot = {slot: [cell_per_round]}."""
    n_rounds = max(len(v) for v in columns_by_slot.values())
    rounds = {}
    for rn in range(1, n_rounds + 1):
        rounds[rn] = [columns_by_slot[slot][rn - 1] for slot in range(1, len(team_names) + 1)]
    return _write_board(tmp_path, team_names=team_names, rounds=rounds)


def test_owner_map_derived_from_overlap_with_shuffled_columns(tmp_path):
    """
    Board column order is shuffled vs ESPN team_id order; the owner map
    must still be derived correctly from roster overlap.
    """
    picks = _four_owner_picks()
    # Column order G2, G1, G4, G3 (shuffled vs ESPN team ids 1,2,3,4)
    path = _board_from_columns(
        tmp_path,
        team_names=["Beta", "Alpha", "Delta", "Gamma"],
        columns_by_slot={
            1: ["Bijan Robinson", "Puka Nacua", "Marvin Harrison", "De'Von Achane"],
            2: ["Christian McCaffrey", "Jahmyr Gibbs", "Garrett Wilson", "Breece Hall"],
            3: ["Saquon Barkley", "Josh Jacobs", "Amari Cooper", "Trevor Lawrence"],
            4: ["Tyreek Hill", "Najee Harris", "Drake London", "Kyle Pitts"],
        },
    )
    board = parse_board(str(path))
    res = match_board_to_picks(board, picks, snake=True)
    assert res.owner_map[1] == "G2"
    assert res.owner_map[2] == "G1"
    assert res.owner_map[3] == "G4"
    assert res.owner_map[4] == "G3"
    assert len(res.placements) == 16
    assert res.match_rate == 1.0


def test_ambiguous_overlap_tie_is_refused(tmp_path):
    """
    When a column's best overlap ties its runner-up, the owner map is
    REFUSED (not guessed) and the result carries an error.
    """
    # G1 and G2 have IDENTICAL rosters -> their two columns tie between
    # them. G3, G4 are distinct so their columns resolve cleanly, but a
    # single ambiguous column refuses the WHOLE map.
    picks = [
        _pick(1, 1, "G1", "Same Player A", display="Alpha", team_id=1),
        _pick(2, 1, "G1", "Same Player B", display="Alpha", team_id=1),
        _pick(3, 1, "G1", "Same Player C", display="Alpha", team_id=1),
        _pick(4, 1, "G1", "Same Player D", display="Alpha", team_id=1),
        _pick(5, 1, "G2", "Same Player A", display="Beta", team_id=2),
        _pick(6, 1, "G2", "Same Player B", display="Beta", team_id=2),
        _pick(7, 1, "G2", "Same Player C", display="Beta", team_id=2),
        _pick(8, 1, "G2", "Same Player D", display="Beta", team_id=2),
        _pick(9, 1, "G3", "Tyreek Hill", display="Gamma", team_id=3),
        _pick(10, 1, "G3", "Najee Harris", display="Gamma", team_id=3),
        _pick(11, 1, "G3", "Drake London", display="Gamma", team_id=3),
        _pick(12, 1, "G3", "Kyle Pitts", display="Gamma", team_id=3),
        _pick(13, 1, "G4", "Saquon Barkley", display="Delta", team_id=4),
        _pick(14, 1, "G4", "Josh Jacobs", display="Delta", team_id=4),
        _pick(15, 1, "G4", "Amari Cooper", display="Delta", team_id=4),
        _pick(16, 1, "G4", "Trevor Lawrence", display="Delta", team_id=4),
    ]
    path = _board_from_columns(
        tmp_path,
        team_names=["Alpha", "Beta", "Gamma", "Delta"],
        columns_by_slot={
            1: ["Same Player A", "Same Player B", "Same Player C", "Same Player D"],
            2: ["Same Player A", "Same Player B", "Same Player C", "Same Player D"],
            3: ["Tyreek Hill", "Najee Harris", "Drake London", "Kyle Pitts"],
            4: ["Saquon Barkley", "Josh Jacobs", "Amari Cooper", "Trevor Lawrence"],
        },
    )
    board = parse_board(str(path))
    res = match_board_to_picks(board, picks, snake=True)
    assert res.owner_map == {}
    assert res.errors
    assert any("ambiguous" in e for e in res.errors)
    assert res.match_rate == 0.0
    assert len(res.unmatched_picks) == 16


def test_too_thin_overlap_below_floor_is_refused(tmp_path):
    """A column whose best overlap is below MIN_OWNER_OVERLAP is refused."""
    picks = _four_owner_picks()
    # Each column only has ONE matching cell (round 1); rounds 2-4 are
    # gibberish -> overlap 1 < MIN_OWNER_OVERLAP=4.
    path = _board_from_columns(
        tmp_path,
        team_names=["Alpha", "Beta", "Gamma", "Delta"],
        columns_by_slot={
            1: ["Christian McCaffrey", "zz gib one", "cc gib", "gg gib"],
            2: ["Bijan Robinson", "qq gib two", "dd gib", "hh gib"],
            3: ["Tyreek Hill", "aa gib", "ee gib", "ii gib"],
            4: ["Saquon Barkley", "bb gib", "ff gib", "jj gib"],
        },
    )
    board = parse_board(str(path))
    res = match_board_to_picks(board, picks, snake=True)
    assert res.owner_map == {}
    assert res.errors


# ---------------------------------------------------------------------------
# nickname resolution
# ---------------------------------------------------------------------------


def test_nickname_resolution_cmc_gibbs_bijan(tmp_path):
    """
    'CMC' -> Christian McCaffrey (acronym), 'J. Gibbs' -> Jahmyr Gibbs
    (last-token / initial-surname), 'Bijan' -> Bijan Robinson (fallback
    token-overlap). Tested within owner G1's column on a 4-team board.
    """
    # G1 owns the nickname targets; other owners hold clean rosters.
    picks = [
        _pick(1, 1, "G1", "Christian McCaffrey"),
        _pick(2, 1, "G1", "Jahmyr Gibbs"),
        _pick(3, 1, "G1", "Bijan Robinson"),
        _pick(4, 1, "G1", "Garrett Wilson"),
        _pick(5, 1, "G2", "Breece Hall"),
        _pick(6, 1, "G2", "Puka Nacua"),
        _pick(7, 1, "G2", "Marvin Harrison"),
        _pick(8, 1, "G2", "De'Von Achane"),
        _pick(9, 1, "G3", "Tyreek Hill"),
        _pick(10, 1, "G3", "Najee Harris"),
        _pick(11, 1, "G3", "Drake London"),
        _pick(12, 1, "G3", "Kyle Pitts"),
        _pick(13, 1, "G4", "Saquon Barkley"),
        _pick(14, 1, "G4", "Josh Jacobs"),
        _pick(15, 1, "G4", "Amari Cooper"),
        _pick(16, 1, "G4", "Trevor Lawrence"),
    ]
    path = _board_from_columns(
        tmp_path,
        team_names=["Alpha", "Beta", "Gamma", "Delta"],
        columns_by_slot={
            # G1's column uses nicknames
            1: ["CMC", "J. Gibbs", "Bijan", "Wilson"],
            2: ["Breece Hall", "Puka Nacua", "Marvin Harrison", "Achane"],
            3: ["Tyreek Hill", "Najee Harris", "Drake London", "Pitts"],
            4: ["Saquon Barkley", "Josh Jacobs", "Amari Cooper", "Lawrence"],
        },
    )
    board = parse_board(str(path))
    res = match_board_to_picks(board, picks, snake=True)
    assert res.match_rate == 1.0
    placed = {pl.cell_text: _name(pl.pick) for pl in res.placements}
    assert placed["CMC"] == "Christian McCaffrey"
    assert placed["J. Gibbs"] == "Jahmyr Gibbs"
    assert placed["Bijan"] == "Bijan Robinson"
    methods = {pl.cell_text: pl.method for pl in res.placements}
    assert methods["CMC"] == "acronym"


def _name(pick):
    return pick.canonical_name or pick.raw_player_name


# ---------------------------------------------------------------------------
# THE REGRESSION TEST THAT MATTERS
# ---------------------------------------------------------------------------


def test_cross_owner_surname_does_not_cross_match(tmp_path):
    """
    Two different players sharing a surname, on two different owners,
    must NOT cross-match. A cell reading the surname must resolve to the
    correct owner's player or be reported unmatched, never to the other
    owner's player of the same surname.

    History: a bare-surname cross-owner match (a cell reading 'Bijan'
    was wrongly matched to a different player surnamed Robinson on
    another team) produced a false positive that this importer must not
    repeat. The within-owner matching constraint prevents it entirely.
    """
    # G1 owns Allen Robinson (a different Robinson); G2 owns Bijan
    # Robinson. G3, G4 own filler players with no surname collisions.
    picks = [
        _pick(1, 1, "G1", "Allen Robinson", display="Alpha", team_id=1),
        _pick(2, 1, "G1", "Stefon Diggs", display="Alpha", team_id=1),
        _pick(3, 1, "G1", "Cooper Kupp", display="Alpha", team_id=1),
        _pick(4, 1, "G1", "Travis Kelce", display="Alpha", team_id=1),
        _pick(5, 1, "G2", "Bijan Robinson", display="Beta", team_id=2),
        _pick(6, 1, "G2", "Justin Jefferson", display="Beta", team_id=2),
        _pick(7, 1, "G2", "Ja'Marr Chase", display="Beta", team_id=2),
        _pick(8, 1, "G2", "Tee Higgins", display="Beta", team_id=2),
        _pick(9, 1, "G3", "Tyreek Hill", display="Gamma", team_id=3),
        _pick(10, 1, "G3", "Najee Harris", display="Gamma", team_id=3),
        _pick(11, 1, "G3", "Drake London", display="Gamma", team_id=3),
        _pick(12, 1, "G3", "Kyle Pitts", display="Gamma", team_id=3),
        _pick(13, 1, "G4", "Saquon Barkley", display="Delta", team_id=4),
        _pick(14, 1, "G4", "Josh Jacobs", display="Delta", team_id=4),
        _pick(15, 1, "G4", "Amari Cooper", display="Delta", team_id=4),
        _pick(16, 1, "G4", "Trevor Lawrence", display="Delta", team_id=4),
    ]
    # Slot 1 (G1) has a "Robinson" cell -> must match Allen Robinson.
    # Slot 2 (G2) has a "Bijan" cell -> must match Bijan Robinson, NEVER
    # Allen Robinson despite the shared surname "Robinson".
    path = _board_from_columns(
        tmp_path,
        team_names=["Alpha", "Beta", "Gamma", "Delta"],
        columns_by_slot={
            1: ["Robinson", "Diggs", "Kupp", "Kelce"],
            2: ["Bijan", "Jefferson", "Chase", "Higgins"],
            3: ["Tyreek Hill", "Najee Harris", "Drake London", "Kyle Pitts"],
            4: ["Saquon Barkley", "Josh Jacobs", "Amari Cooper", "Lawrence"],
        },
    )
    board = parse_board(str(path))
    res = match_board_to_picks(board, picks, snake=True)

    assert res.owner_map[1] == "G1"
    assert res.owner_map[2] == "G2"

    # Each placement's pick owner must equal the slot's assigned owner:
    # no pick ever crosses to another owner.
    for pl in res.placements:
        assert pl.pick.member_guid == res.owner_map[pl.slot], (
            f"placement {pl.cell_text!r} on slot {pl.slot} (owner "
            f"{res.owner_map[pl.slot]}) consumed a pick from guid "
            f"{pl.pick.member_guid} -> cross-owner leak"
        )

    # Specifically: "Bijan" on slot 2 must have matched G2's Bijan
    # Robinson, not G1's Allen Robinson.
    bijan = next(pl for pl in res.placements if pl.cell_text == "Bijan")
    assert bijan.pick.canonical_name == "Bijan Robinson"
    assert bijan.pick.member_guid == "G2"


# ---------------------------------------------------------------------------
# consume-once
# ---------------------------------------------------------------------------


def test_each_espn_pick_consumed_at_most_once(tmp_path):
    """
    Two cells that resolve to the same player name: only the first
    matches; the second is reported unmatched (pick consumed once).
    """
    picks = _four_owner_picks()
    # Slot 1 (G1) has a duplicate CMC cell -> second must be unmatched.
    path = _board_from_columns(
        tmp_path,
        team_names=["Alpha", "Beta", "Gamma", "Delta"],
        columns_by_slot={
            1: ["Christian McCaffrey", "Christian McCaffrey", "Jahmyr Gibbs", "Garrett Wilson"],
            2: ["Bijan Robinson", "Puka Nacua", "Marvin Harrison", "Achane"],
            3: ["Tyreek Hill", "Najee Harris", "Drake London", "Pitts"],
            4: ["Saquon Barkley", "Josh Jacobs", "Amari Cooper", "Lawrence"],
        },
    )
    board = parse_board(str(path))
    res = match_board_to_picks(board, picks, snake=True)
    # 15 placements (one CMC cell is unmatched); 16 cells total.
    assert len(res.placements) == 15
    assert len(res.unmatched_cells) == 1
    assert res.unmatched_cells[0].text == "Christian McCaffrey"
    placed_ids = [id(pl.pick) for pl in res.placements]
    assert len(placed_ids) == len(set(placed_ids))


# ---------------------------------------------------------------------------
# apply_board with a fake engine
# ---------------------------------------------------------------------------


class FakeEngine:
    """
    Minimal in-memory stand-in for the ODMantic engine: just enough to
    back apply_board's find/save_all calls without a Mongo.
    """

    def __init__(self, picks):
        self._picks = list(picks)

    async def find(self, model, query=None):
        return list(self._picks)

    async def save_all(self, objs):
        for o in objs:
            for i, p in enumerate(self._picks):
                if p is o:
                    self._picks[i] = o


def _good_board_path(tmp_path):
    """A 4-team board + picks that match cleanly at 100% match_rate."""
    picks = _four_owner_picks()
    path = _board_from_columns(
        tmp_path,
        team_names=["Alpha", "Beta", "Gamma", "Delta"],
        columns_by_slot={
            1: ["Christian McCaffrey", "Jahmyr Gibbs", "Garrett Wilson", "Breece Hall"],
            2: ["Bijan Robinson", "Puka Nacua", "Marvin Harrison", "De'Von Achane"],
            3: ["Tyreek Hill", "Najee Harris", "Drake London", "Kyle Pitts"],
            4: ["Saquon Barkley", "Josh Jacobs", "Amari Cooper", "Trevor Lawrence"],
        },
    )
    return path, picks


def test_apply_board_dry_run_makes_no_mutation(tmp_path):
    path, picks = _good_board_path(tmp_path)
    before = [
        (p.overall_pick, p.round_num, p.round_pick, p.draft_order_verified)
        for p in picks
    ]
    engine = FakeEngine(picks)
    report = asyncio.run(
        apply_board(
            engine,
            espn_league_id=1,
            season=2024,
            board_path=str(path),
            dry_run=True,
        )
    )
    after = [
        (p.overall_pick, p.round_num, p.round_pick, p.draft_order_verified)
        for p in picks
    ]
    assert before == after
    assert report["dry_run"] is True
    assert report["written"] is False
    assert report["match_rate"] == 1.0


def test_apply_board_writes_when_applied(tmp_path):
    path, picks = _good_board_path(tmp_path)
    engine = FakeEngine(picks)
    report = asyncio.run(
        apply_board(
            engine,
            espn_league_id=1,
            season=2024,
            board_path=str(path),
            dry_run=False,
        )
    )
    assert report["written"] is True
    assert all(p.draft_order_verified for p in picks)
    overalls = [p.overall_pick for p in picks]
    assert len(set(overalls)) == len(overalls)


def test_apply_board_match_rate_guard_blocks_bad_board(tmp_path):
    """A board with low match_rate is refused unless force=True."""
    picks = _four_owner_picks()
    # Rounds 1-4 match cleanly (owner map derives); rounds 5-8 are
    # gibberish so match_rate = 16/32 = 0.5 < MIN_MATCH_RATE.
    path = _board_from_columns(
        tmp_path,
        team_names=["Alpha", "Beta", "Gamma", "Delta"],
        columns_by_slot={
            1: ["Christian McCaffrey", "Jahmyr Gibbs", "Garrett Wilson", "Breece Hall",
                "zz gib one", "cc gib", "gg gib", "kk gib"],
            2: ["Bijan Robinson", "Puka Nacua", "Marvin Harrison", "De'Von Achane",
                "qq gib two", "dd gib", "hh gib", "ll gib"],
            3: ["Tyreek Hill", "Najee Harris", "Drake London", "Kyle Pitts",
                "aa gib", "ee gib", "ii gib", "mm gib"],
            4: ["Saquon Barkley", "Josh Jacobs", "Amari Cooper", "Trevor Lawrence",
                "bb gib", "ff gib", "jj gib", "nn gib"],
        },
    )
    engine = FakeEngine(picks)
    report = asyncio.run(
        apply_board(
            engine,
            espn_league_id=1,
            season=2024,
            board_path=str(path),
            dry_run=False,
        )
    )
    assert report["refused"] is True
    assert any("match_rate" in r for r in report["refusal_reasons"])
    assert report["written"] is False
    assert not any(p.draft_order_verified for p in picks)

    # force=True overrides the guard and writes the matched subset
    engine2 = FakeEngine(picks)
    report2 = asyncio.run(
        apply_board(
            engine2,
            espn_league_id=1,
            season=2024,
            board_path=str(path),
            dry_run=False,
            force=True,
        )
    )
    assert report2["written"] is True
    matched = [p for p in picks if p.draft_order_verified]
    assert len(matched) == 16


def test_apply_board_duplicate_overall_pick_guard(tmp_path):
    """
    The duplicate-overall_pick guard is a SAFETY net: under normal
    operation parse_board collapses same-(round,slot) cells into one
    dict key, so duplicates cannot arise. We exercise the guard by
    monkeypatching overall_pick_for to return a constant, simulating a
    corrupt board.
    """
    path, picks = _good_board_path(tmp_path)
    import data_sources.draft_board_import as mod

    orig = mod.overall_pick_for
    mod.overall_pick_for = lambda *a, **k: 1  # force collision
    try:
        engine = FakeEngine(picks)
        report = asyncio.run(
            apply_board(
                engine,
                espn_league_id=1,
                season=2024,
                board_path=str(path),
                dry_run=False,
            )
        )
        assert report["refused"] is True
        assert any(
            "duplicate overall_pick" in r for r in report["refusal_reasons"]
        )
        assert report["written"] is False
        # force overrides
        engine2 = FakeEngine(picks)
        report2 = asyncio.run(
            apply_board(
                engine2,
                espn_league_id=1,
                season=2024,
                board_path=str(path),
                dry_run=False,
                force=True,
            )
        )
        assert report2["written"] is True
    finally:
        mod.overall_pick_for = orig
